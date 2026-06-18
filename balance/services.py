"""
services.py - لایه منطق کسب‌وکار برای سیستم موازنه متریال جهانپارس
=======================================================================
این فایل مسئول پردازش داده‌ها، اجرای فرمول‌های موازنه و تولید خروجی اکسل است.
ساختار خروجی:
  - یک شیت به ازای هر پیمانکار
  - بالای هر شیت: اطلاعات پیمانکار، شماره قرارداد، موضوع قرارداد، بازه زمانی
  - جدول داده‌ها با قابلیت فیلتر (AutoFilter)
"""

import io
from decimal import Decimal, ROUND_HALF_UP
import datetime
import jdatetime

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from django.db.models import Sum, Min, Max


# ─────────────────────────────────────────────────────────────────────────────
# ثابت‌های رنگ برای قالب‌بندی اکسل
# ─────────────────────────────────────────────────────────────────────────────
COLOR_HEADER_BG      = "1F3864"   # آبی تیره - هدر اصلی
COLOR_HEADER_FONT    = "FFFFFF"   # سفید
COLOR_SUB_HEADER_BG  = "2E75B6"   # آبی متوسط - ردیف عناوین ستون
COLOR_INFO_BG        = "EBF3FB"   # آبی خیلی روشن - بلوک اطلاعات پیمانکار
COLOR_INFO_LABEL_BG  = "D6E4F0"   # کمی تیره‌تر برای برچسب‌ها
COLOR_ROW_ODD        = "DCE6F1"   # آبی خیلی روشن - ردیف‌های فرد
COLOR_ROW_EVEN       = "FFFFFF"   # سفید - ردیف‌های زوج
COLOR_POSITIVE       = "C6EFCE"   # سبز روشن - مازاد پرداخت
COLOR_POSITIVE_FONT  = "276221"   # سبز تیره
COLOR_NEGATIVE       = "FFC7CE"   # قرمز روشن - کسری متریال
COLOR_NEGATIVE_FONT  = "9C0006"   # قرمز تیره
COLOR_ZERO           = "FFEB9C"   # زرد روشن - ایده‌آل
COLOR_ZERO_FONT      = "9C6500"   # نارنجی تیره
COLOR_BORDER         = "BDD7EE"   # آبی کم‌رنگ برای خطوط جدول
COLOR_INFO_BORDER    = "AEC6CF"   # رنگ حاشیه بلوک اطلاعات


# ─────────────────────────────────────────────────────────────────────────────
# کش استایل‌های OpenPyXL (افزایش بسیار زیاد سرعت تولید فایل)
# ─────────────────────────────────────────────────────────────────────────────
GLOBAL_BORDER = Border(
    left=Side(style='thin', color=COLOR_BORDER),
    right=Side(style='thin', color=COLOR_BORDER),
    top=Side(style='thin', color=COLOR_BORDER),
    bottom=Side(style='thin', color=COLOR_BORDER)
)

GLOBAL_ALIGNMENT_RIGHT = Alignment(horizontal='right', vertical='center', readingOrder=2)
GLOBAL_ALIGNMENT_CENTER = Alignment(horizontal='center', vertical='center', readingOrder=2)
GLOBAL_ALIGNMENT_RIGHT_WRAP = Alignment(horizontal='right', vertical='center', wrap_text=True, readingOrder=2)
GLOBAL_ALIGNMENT_CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True, readingOrder=2)

FONT_DATA = Font(name='Calibri', size=10, bold=False)
FONT_DATA_BOLD = Font(name='Calibri', size=10, bold=True)
FONT_POS = Font(name='Calibri', size=10, bold=True, color=COLOR_POSITIVE_FONT)
FONT_NEG = Font(name='Calibri', size=10, bold=True, color=COLOR_NEGATIVE_FONT)
FONT_ZERO = Font(name='Calibri', size=10, bold=True, color=COLOR_ZERO_FONT)

FILL_ODD = PatternFill(fill_type='solid', fgColor=COLOR_ROW_ODD)
FILL_EVEN = PatternFill(fill_type='solid', fgColor=COLOR_ROW_EVEN)
FILL_POS = PatternFill(fill_type='solid', fgColor=COLOR_POSITIVE)
FILL_NEG = PatternFill(fill_type='solid', fgColor=COLOR_NEGATIVE)
FILL_ZERO = PatternFill(fill_type='solid', fgColor=COLOR_ZERO)


def _make_border(color=COLOR_BORDER):
    """ایجاد یک شیء Border (فقط برای هدرها یا بخش‌های خاص استفاده شود)"""
    side = Side(style='thin', color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _rtl_alignment(horizontal='right', vertical='center', wrap=False):
    """تنظیم راست‌چین RTL برای سلول‌های فارسی (فقط هدرها)"""
    return Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap,
        readingOrder=2,  # 2 = RTL
    )


def _apply_header_style(cell, text, font_size=11, bold=True, bg_color=COLOR_HEADER_BG, font_color=COLOR_HEADER_FONT):
    """اعمال استایل هدر به یک سلول."""
    cell.value = text
    cell.font = Font(name='Calibri', bold=bold, size=font_size, color=font_color)
    cell.fill = PatternFill(fill_type='solid', fgColor=bg_color)
    cell.alignment = _rtl_alignment(horizontal='center')
    cell.border = _make_border(COLOR_BORDER)


from openpyxl.styles import NamedStyle

def _register_named_styles(wb):
    """ثبت استایل‌های کش شده برای پرهیز از هشینگ مجدد در openpyxl"""
    if 'data_even_right_norm_str' in wb.named_styles:
        return  # قبلا ثبت شده
        
    styles_to_add = []
    
    # 1. Data Styles
    for row_type in ('odd', 'even'):
        for align in ('right', 'center'):
            for bold in (True, False):
                for num_key, num_fmt in [('str', None), ('money', '#,##0.00'), ('pct', '0.00')]:
                    name = f"data_{row_type}_{align}_{'bold' if bold else 'norm'}_{num_key}"
                    ns = NamedStyle(name=name)
                    ns.font = FONT_DATA_BOLD if bold else FONT_DATA
                    ns.fill = FILL_ODD if row_type == 'odd' else FILL_EVEN
                    ns.alignment = GLOBAL_ALIGNMENT_CENTER if align == 'center' else GLOBAL_ALIGNMENT_RIGHT
                    ns.border = GLOBAL_BORDER
                    if num_fmt:
                        ns.number_format = num_fmt
                    styles_to_add.append(ns)

    # 2. Balance Styles
    for bal in ('pos', 'neg', 'zero', 'review'):
        name = f"bal_{bal}"
        ns = NamedStyle(name=name)
        if bal == 'pos':
            ns.font = FONT_POS; ns.fill = FILL_POS
        elif bal == 'neg':
            ns.font = FONT_NEG; ns.fill = FILL_NEG
        elif bal == 'zero':
            ns.font = FONT_ZERO; ns.fill = FILL_ZERO
        else: # review
            ns.font = Font(name='Calibri', size=10, bold=True, color="595959")
            ns.fill = PatternFill(fill_type='solid', fgColor="F2F2F2")
        ns.alignment = GLOBAL_ALIGNMENT_CENTER
        ns.border = GLOBAL_BORDER
        if bal != 'review':
            ns.number_format = '#,##0.00'
        styles_to_add.append(ns)
        
    # 3. Transaction Type Styles
    for txn_type in ('IN', 'OUT'):
        name = f"data_{txn_type.lower()}"
        ns = NamedStyle(name=name)
        if txn_type == 'IN':
            ns.font = FONT_POS; ns.fill = FILL_POS
        else:
            ns.font = FONT_NEG; ns.fill = FILL_NEG
        ns.alignment = GLOBAL_ALIGNMENT_CENTER
        ns.border = GLOBAL_BORDER
        styles_to_add.append(ns)
        
    for s in styles_to_add:
        wb.add_named_style(s)


def _apply_data_style(cell, value, row_idx, number_format=None, bold=False, wrap=False, horizontal='right'):
    """اعمال استایل داده به یک سلول معمولی (نسخه NamedStyle بسیار سریع)."""
    cell.value = value
    
    row_type = 'odd' if row_idx % 2 == 0 else 'even'
    is_bold = 'bold' if bold else 'norm'
    
    if number_format == '#,##0.00':
        num_key = 'money'
    elif number_format == '0.00':
        num_key = 'pct'
    elif number_format:
        num_key = 'str'  # Fallback
    else:
        num_key = 'str'
        
    style_name = f"data_{row_type}_{horizontal}_{is_bold}_{num_key}"
    cell.style = style_name
    
    # Override number_format if it is a custom one not in our standard map
    if number_format and num_key == 'str':
        cell.number_format = number_format
        
    # Wrap text override if needed
    if wrap:
        cell.alignment = GLOBAL_ALIGNMENT_CENTER_WRAP if horizontal == 'center' else GLOBAL_ALIGNMENT_RIGHT_WRAP


def _apply_balance_style(cell, value):
    """اعمال استایل شرطی به ستون موازنه (نسخه NamedStyle بسیار سریع)"""
    if isinstance(value, str) and "در دست بررسی" in value:
        cell.value = value
        cell.style = 'bal_review'
        return
        
    cell.value = float(value)
    if value > 0:
        cell.style = 'bal_pos'
    elif value < 0:
        cell.style = 'bal_neg'
    else:
        cell.style = 'bal_zero'


def _balance_label(value):
    """تبدیل عدد موازنه به برچسب توصیفی فارسی."""
    if isinstance(value, str) and "در دست بررسی" in value:
        return "در دست بررسی ⏳"
    if value > 0:
        return "مازاد پرداخت ✔"
    elif value < 0:
        return "کسری متریال ✘"
    else:
        return "ایده‌آل ✔"


def _write_info_cell(ws, row, col_start, col_end, text, is_label=False):
    """
    نوشتن و استایل‌دهی یک سلول اطلاعاتی در مشخصات بالا.
    """
    if col_start == col_end:
        cell = ws.cell(row=row, column=col_start)
    else:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        cell = ws.cell(row=row, column=col_start)
    
    cell.value = text
    cell.alignment = _rtl_alignment(horizontal='right')
    
    if is_label:
        cell.font = Font(name='Calibri', bold=True, size=10, color="1F3864")
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_LABEL_BG)
    else:
        cell.font = Font(name='Calibri', size=10, color="333333")
        cell.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_BG)
        
    cell.border = _make_border(COLOR_INFO_BORDER)
    
    # برای سلول‌های merge شده باید حاشیه و بک‌گراند را به کل ستون‌ها اعمال کنیم تا زشت نشود
    for r in range(row, row + 1):
        for c in range(col_start, col_end + 1):
            cell_in_merge = ws.cell(row=r, column=c)
            cell_in_merge.border = _make_border(COLOR_INFO_BORDER)
            if not is_label:
                cell_in_merge.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_BG)
            else:
                cell_in_merge.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_LABEL_BG)


def _build_contractor_sheet(ws, contractor, rows):
    """
    ساخت یک شیت کامل برای یک پیمانکار.
    """
    ws.sheet_view.rightToLeft = True  # RTL برای کل شیت
    TOTAL_COLS = 15

    # ── ۱. ردیف عنوان اصلی ─────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    title_cell = ws.cell(row=1, column=1)
    _apply_header_style(
        title_cell,
        "گزارش موازنه متریال - شرکت جهانپارس",
        font_size=16,
        bg_color=COLOR_HEADER_BG,
    )
    ws.row_dimensions[1].height = 38

    # ── ۲. ردیف توضیح فرمول ────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = "موازنه = کل تحویلی − (کار تاییدشده + پرتی مجاز)    |    سبز: مازاد پرداخت    |    قرمز: کسری متریال    |    زرد: ایده‌آل"
    sub_cell.font  = Font(name='Calibri', size=9, italic=True, color="4472C4")
    sub_cell.fill  = PatternFill(fill_type='solid', fgColor="EBF3FB")
    sub_cell.alignment = _rtl_alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    # ── ۳. بلوک اطلاعات پیمانکار (ردیف ۳) ────────────────────────
    contractor_name = contractor.get_full_name()

    # ردیف ۳: نام پیمانکار، تاریخ گزارش
    _write_info_cell(ws, 3, 1, 1, "نام پیمانکار :", is_label=True)
    _write_info_cell(ws, 3, 2, 8, contractor_name, is_label=False)
    
    _write_info_cell(ws, 3, 9, 10, "تاریخ گزارش :", is_label=True)
    _write_info_cell(ws, 3, 11, 15, str(jdatetime.date.today()), is_label=False)
    
    ws.row_dimensions[3].height = 20

    # ── ۴. خط جداکننده ردیف ۴ ──────────────────────────────────────────────
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=TOTAL_COLS)
    sep = ws.cell(row=4, column=1)
    sep.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
    ws.row_dimensions[4].height = 6

    # ── ۵. هدر ستون‌های جدول (ردیف ۵) ─────────────────────────────────────
    HEADERS = [
        "ردیف",
        "شماره قرارداد",
        "موضوع قرارداد",
        "رسته کاری",
        "نام کالا",
        "سایز",
        "جنس",
        "ضخامت",
        "واحد",
        "کل متریال\nتحویلی",
        "مقدار کار\nتاییدشده",
        "درصد\nپرتی (%)",
        "پرتی\nمجاز",
        "موازنه\n(انحراف)",
        "وضعیت نهایی",
    ]
    COL_WIDTHS = [6, 16, 25, 16, 28, 12, 14, 12, 8, 14, 14, 10, 12, 14, 20]

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=5, column=col_idx)
        _apply_header_style(cell, header, font_size=10, bg_color=COLOR_SUB_HEADER_BG)
        cell.alignment = _rtl_alignment(horizontal='center', wrap=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[5].height = 38

    # ── ۶. ردیف‌های داده (از ردیف ۶) ─────────────────────────────────────
    NUM_FMT = '#,##0.00'

    for row_offset, row_data in enumerate(rows):
        excel_row = row_offset + 6

        _apply_data_style(ws.cell(excel_row, 1),  row_offset + 1,                    row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 2),  row_data['contract_number'],        row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 3),  row_data['contract_subject'],       row_offset)
        _apply_data_style(ws.cell(excel_row, 4),  row_data['work_category'],          row_offset)
        _apply_data_style(ws.cell(excel_row, 5),  row_data['material_name'],          row_offset)
        _apply_data_style(ws.cell(excel_row, 6),  row_data['size'],                   row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 7),  row_data['mat_type'],               row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 8),  row_data['thickness'],              row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 9),  row_data['unit'],                   row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 10), float(row_data['total_issued']),    row_offset, number_format=NUM_FMT)
        _apply_data_style(ws.cell(excel_row, 11), float(row_data['approved_work']),   row_offset, number_format=NUM_FMT)
        _apply_data_style(ws.cell(excel_row, 12), row_data['waste_pct'],              row_offset, number_format='0.00')
        _apply_data_style(ws.cell(excel_row, 13), float(row_data['allowed_waste']),   row_offset, number_format=NUM_FMT)
        _apply_balance_style(ws.cell(excel_row, 14), row_data['balance'])

        # وضعیت نهایی
        status_cell = ws.cell(excel_row, 15)
        status_cell.value = row_data['balance_label']
        if isinstance(row_data['balance'], str):
            status_cell.style = 'bal_review'
        elif row_data['balance'] > 0:
            status_cell.style = 'bal_pos'
        elif row_data['balance'] < 0:
            status_cell.style = 'bal_neg'
        else:
            status_cell.style = 'bal_zero'
            
        ws.row_dimensions[excel_row].height = 22

    # ── ۷. ردیف جمع کل ─────────────────────────────────────────────────────
    if rows:
        total_row = len(rows) + 6
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=9)
        _apply_header_style(ws.cell(total_row, 1), "جمع کل", font_size=11, bg_color=COLOR_HEADER_BG)

        grand_issued   = sum(r['total_issued']  for r in rows)
        grand_approved = sum(r['approved_work'] for r in rows)
        grand_waste    = sum(r['allowed_waste'] for r in rows)
        grand_balance  = sum(r['balance']       for r in rows if not isinstance(r['balance'], str))

        for col, val in [(10, grand_issued), (11, grand_approved), (13, grand_waste)]:
            c = ws.cell(total_row, col)
            _apply_header_style(c, float(val), bg_color=COLOR_HEADER_BG)
            c.number_format = NUM_FMT

        # ستون درصد پرتی در ردیف جمع خالی ولی با رنگ
        ws.cell(total_row, 12).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        ws.cell(total_row, 12).border = GLOBAL_BORDER

        _apply_balance_style(ws.cell(total_row, 14), grand_balance)
        ws.cell(total_row, 14).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        ws.cell(total_row, 14).font = Font(name='Calibri', size=11, bold=True, color=COLOR_HEADER_FONT)

        ws.cell(total_row, 15).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        ws.cell(total_row, 15).border = GLOBAL_BORDER
        ws.row_dimensions[total_row].height = 26

        # ── ۸. AutoFilter روی جدول داده ─────────────────────────────────────
        last_data_row = total_row - 1
        first_col_letter = get_column_letter(1)
        last_col_letter = get_column_letter(TOTAL_COLS)
        ws.auto_filter.ref = f"{first_col_letter}5:{last_col_letter}{last_data_row}"

    # ── ۹. فریز پنل (هدر و اطلاعات پیمانکار ثابت بماند) ────────────────────
    ws.freeze_panes = "A6"
    ws.print_title_rows = '1:5'
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1


def _build_global_sheet(ws, rows):
    """
    ساخت شیت گزارش موازنه کل کارگاه.
    """
    ws.sheet_view.rightToLeft = True  # RTL برای کل شیت
    TOTAL_COLS = 16

    # ── ۱. ردیف عنوان اصلی ─────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    title_cell = ws.cell(row=1, column=1)
    _apply_header_style(
        title_cell,
        "گزارش موازنه متریال کل - شرکت جهانپارس",
        font_size=16,
        bg_color=COLOR_HEADER_BG,
    )
    ws.row_dimensions[1].height = 38

    # ── ۲. ردیف توضیح فرمول ────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = "موازنه = کل تحویلی − (کار تاییدشده + پرتی مجاز)    |    سبز: مازاد پرداخت    |    قرمز: کسری متریال    |    زرد: ایده‌آل"
    sub_cell.font  = Font(name='Calibri', size=9, italic=True, color="4472C4")
    sub_cell.fill  = PatternFill(fill_type='solid', fgColor="EBF3FB")
    sub_cell.alignment = _rtl_alignment(horizontal='center')
    ws.row_dimensions[2].height = 18

    # ── ۳. بلوک اطلاعات گزارش (ردیف ۳ و ۴ - به صورت فشرده) ────────────────────────
    # ردیف ۳: تاریخ گزارش و توضیح
    _write_info_cell(ws, 3, 1, 2, "تاریخ گزارش :", is_label=True)
    _write_info_cell(ws, 3, 3, 15, str(jdatetime.date.today()), is_label=False)
    ws.row_dimensions[3].height = 20

    # ردیف ۴: حوزه گزارش
    _write_info_cell(ws, 4, 1, 2, "حوزه گزارش :", is_label=True)
    _write_info_cell(ws, 4, 3, 15, "موازنه کل (به تفکیک تمامی پیمانکاران)", is_label=False)
    ws.row_dimensions[4].height = 20

    # ── ۴. خط جداکننده ردیف ۵ ──────────────────────────────────────────────
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=TOTAL_COLS)
    sep = ws.cell(row=5, column=1)
    sep.fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
    ws.row_dimensions[5].height = 6

    # ── ۵. هدر ستون‌های جدول (ردیف ۶) ─────────────────────────────────────
    HEADERS = [
        "ردیف",
        "پیمانکار",
        "شماره قرارداد",
        "موضوع قرارداد",
        "رسته کاری",
        "نام کالا",
        "سایز",
        "جنس",
        "ضخامت",
        "واحد",
        "کل متریال\nتحویلی",
        "مقدار کار\nتاییدشده",
        "درصد\nپرتی (%)",
        "پرتی\nمجاز",
        "موازنه\n(انحراف)",
        "وضعیت نهایی",
    ]
    COL_WIDTHS = [6, 22, 16, 25, 16, 24, 12, 14, 12, 8, 14, 14, 10, 12, 14, 18]

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=6, column=col_idx)
        _apply_header_style(cell, header, font_size=10, bg_color=COLOR_SUB_HEADER_BG)
        cell.alignment = _rtl_alignment(horizontal='center', wrap=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[6].height = 38

    # ── ۶. ردیف‌های داده (از ردیف ۷) ─────────────────────────────────────
    NUM_FMT = '#,##0.00'

    for row_offset, row_data in enumerate(rows):
        excel_row = row_offset + 7

        _apply_data_style(ws.cell(excel_row, 1),  row_offset + 1,                    row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 2),  row_data['contractor_name'],       row_offset)
        _apply_data_style(ws.cell(excel_row, 3),  row_data['contract_number'],       row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 4),  row_data['contract_subject'],      row_offset)
        _apply_data_style(ws.cell(excel_row, 5),  row_data['work_category'],         row_offset)
        _apply_data_style(ws.cell(excel_row, 6),  row_data['material_name'],         row_offset)
        _apply_data_style(ws.cell(excel_row, 7),  row_data['size'],                  row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 8),  row_data['mat_type'],              row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 9),  row_data['thickness'],             row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 10), row_data['unit'],                  row_offset, horizontal='center')
        _apply_data_style(ws.cell(excel_row, 11), float(row_data['total_issued']),   row_offset, number_format=NUM_FMT)
        _apply_data_style(ws.cell(excel_row, 12), float(row_data['approved_work']),  row_offset, number_format=NUM_FMT)
        _apply_data_style(ws.cell(excel_row, 13), row_data['waste_pct'],             row_offset, number_format='0.00')
        _apply_data_style(ws.cell(excel_row, 14), float(row_data['allowed_waste']),  row_offset, number_format=NUM_FMT)
        _apply_balance_style(ws.cell(excel_row, 15), row_data['balance'])

        # وضعیت نهایی
        status_cell = ws.cell(excel_row, 16)
        status_cell.value = row_data['balance_label']
        if isinstance(row_data['balance'], str):
            status_cell.style = 'bal_review'
        elif row_data['balance'] > 0:
            status_cell.style = 'bal_pos'
        elif row_data['balance'] < 0:
            status_cell.style = 'bal_neg'
        else:
            status_cell.style = 'bal_zero'
            
        ws.row_dimensions[excel_row].height = 22

    # ── ۷. ردیف جمع کل ─────────────────────────────────────────────────────
    if rows:
        total_row = len(rows) + 7
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=10)
        _apply_header_style(ws.cell(total_row, 1), "جمع کل", font_size=11, bg_color=COLOR_HEADER_BG)

        grand_issued   = sum(r['total_issued']  for r in rows)
        grand_approved = sum(r['approved_work'] for r in rows)
        grand_waste    = sum(r['allowed_waste'] for r in rows)
        grand_balance  = sum(r['balance']       for r in rows if not isinstance(r['balance'], str))

        for col, val in [(11, grand_issued), (12, grand_approved), (14, grand_waste)]:
            c = ws.cell(total_row, col)
            _apply_header_style(c, float(val), bg_color=COLOR_HEADER_BG)
            c.number_format = NUM_FMT

        # ستون درصد پرتی در ردیف جمع خالی ولی با رنگ
        ws.cell(total_row, 13).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        ws.cell(total_row, 13).border = _make_border()

        _apply_balance_style(ws.cell(total_row, 15), grand_balance)
        ws.cell(total_row, 15).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        ws.cell(total_row, 15).font = Font(name='Calibri', size=11, bold=True, color=COLOR_HEADER_FONT)

        ws.cell(total_row, 16).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
        ws.cell(total_row, 16).border = _make_border()
        ws.row_dimensions[total_row].height = 26

        # ── ۸. AutoFilter روی جدول داده ─────────────────────────────────────
        last_data_row = total_row - 1
        first_col_letter = get_column_letter(1)
        last_col_letter = get_column_letter(TOTAL_COLS)
        ws.auto_filter.ref = f"{first_col_letter}6:{last_col_letter}{last_data_row}"

    # ── ۹. فریز پنل (هدر و اطلاعات ثابت بماند) ────────────────────
    ws.freeze_panes = "A7"
    ws.print_title_rows = '1:6'
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1


# ─────────────────────────────────────────────────────────────────────────────
# تابع اصلی تولید گزارش
# ─────────────────────────────────────────────────────────────────────────────
def generate_material_balance_excel(
    contractor_id: int | None = None,
    material_id: int | None = None,
    is_superuser: bool = False,
) -> bytes:
    """
    تولید گزارش موازنه متریال در قالب فایل اکسل.

    آرگومان‌های اختیاری:
        contractor_id: در صورت ارائه، فقط برای آن پیمانکار فیلتر می‌شود.
        material_id:   در صورت ارائه، فقط برای آن متریال فیلتر می‌شود.

    خروجی:
        bytes - محتوای فایل xlsx آماده دانلود یا ذخیره.
        یک شیت جداگانه به ازای هر پیمانکار.

    فرمول‌های محاسباتی:
        ۱. کل متریال تحویلی  = جمع OUT از WarehouseTransaction
        ۲. مقدار کار تاییدشده = جمع approved_quantity از TechnicalOfficeApproval
        ۳. پرتی مجاز         = مقدار کار تاییدشده × (waste_percentage / 100)
        ۴. موازنه            = کل تحویلی - (تاییدشده + پرتی مجاز)
    """
    # import داخلی برای جلوگیری از مشکلات circular import در Django
    from .models import WarehouseTransaction, TechnicalOfficeApproval, MaterialItem, Contractor

    # ─── ۱. بارگذاری و جمع‌بندی داده‌های تراکنش خروجی انبار ───────────────────────────
    issue_qs = WarehouseTransaction.objects.filter(transaction_type='OUT')
    if contractor_id:
        issue_qs = issue_qs.filter(contractor_id=contractor_id)
    if material_id:
        issue_qs = issue_qs.filter(material_id=material_id)

    # جمع‌بندی مقادیر با دیتابیس (سرعت هزار برابری)
    issued_aggs = issue_qs.values('contractor_id', 'material_id', 'contract_number', 'contract_subject').annotate(total_qty=Sum('quantity'))
    issued_map: dict[tuple, Decimal] = {}
    for agg in issued_aggs:
        k = (agg['contractor_id'], agg['material_id'], agg['contract_number'] or '', agg['contract_subject'] or '')
        issued_map[k] = agg['total_qty'] or Decimal('0')

    # ─── ۲. بارگذاری تاییدیه‌های دفتر فنی ─────────────────────────────────
    approval_qs = TechnicalOfficeApproval.objects.all()
    if contractor_id:
        approval_qs = approval_qs.filter(contractor_id=contractor_id)
    if material_id:
        approval_qs = approval_qs.filter(material_id=material_id)

    appr_aggs = approval_qs.values('contractor_id', 'material_id', 'contract_number', 'contract_subject').annotate(total_appr=Sum('approved_quantity'))
    approved_map: dict[tuple, Decimal] = {}
    for agg in appr_aggs:
        k = (agg['contractor_id'], agg['material_id'], agg['contract_number'] or '', agg['contract_subject'] or '')
        approved_map[k] = agg['total_appr'] or Decimal('0')

    # ─── ۳. جمع‌آوری تمام جفت‌های (contractor, material) یکتا ─────────────
    all_keys = set(issued_map.keys()) | set(approved_map.keys())

    if not all_keys:
        # اگر هیچ داده‌ای وجود نداشت، یک فایل خالی با پیام برگردان
        wb = Workbook()
        ws = wb.active
        ws.title = "بدون داده"
        ws.sheet_view.rightToLeft = True
        ws.merge_cells('A1:E1')
        c = ws['A1']
        c.value = "هیچ داده‌ای برای گزارش‌گیری وجود ندارد."
        c.font = Font(name='Calibri', size=12, bold=True, color="9C0006")
        c.alignment = _rtl_alignment(horizontal='center')
        if not is_superuser:
            ws.protection.sheet = True
            ws.protection.password = "jahanpars2026"
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # بارگذاری آبجکت‌های Contractor و MaterialItem
    contractor_ids = {k[0] for k in all_keys if k[0] is not None}
    material_ids   = {k[1] for k in all_keys}

    contractors = {u.id: u for u in Contractor.objects.filter(id__in=contractor_ids)}
    materials   = {m.id: m for m in MaterialItem.objects.select_related('work_category').filter(id__in=material_ids)}

    # ─── ۴. گروه‌بندی ردیف‌ها بر اساس پیمانکار ─────────────────────────────
    # ساختار: {contractor_id: [row_dict, ...]}
    contractor_rows: dict[int, list] = {c_id: [] for c_id in contractor_ids}

    for (c_id, m_id, contract_num, contract_subj) in sorted(all_keys, key=lambda k: (k[0] or 0, k[1], k[2], k[3])):
        material = materials.get(m_id)
        if not material:
            continue

        total_issued  = issued_map.get((c_id, m_id, contract_num, contract_subj), Decimal('0'))
        
        # بررسی اینکه آیا اصلا تاییدیه‌ای برای این قرارداد و متریال وجود دارد یا خیر
        if (c_id, m_id, contract_num, contract_subj) not in approved_map:
            approved_work = Decimal('0')
            allowed_waste = Decimal('0')
            balance = "در دست بررسی برای تایید دفتر فنی"
        else:
            approved_work = approved_map[(c_id, m_id, contract_num, contract_subj)]
            waste_pct     = material.waste_percentage / Decimal('100')
            allowed_waste = (approved_work * waste_pct).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            balance       = (total_issued - (approved_work + allowed_waste)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        row = {
            'contract_number': contract_num or "—",
            'contract_subject': contract_subj or "—",
            'work_category':  material.work_category.name if material.work_category else "—",
            'material_name':  material.name,
            'size':           material.size or "—",
            'mat_type':       material.material_type or "—",
            'thickness':      material.thickness or "—",
            'unit':           material.get_unit_display(),
            'total_issued':   total_issued,
            'approved_work':  approved_work,
            'waste_pct':      float(material.waste_percentage),
            'allowed_waste':  allowed_waste,
            'balance':        balance,
            'balance_label':  _balance_label(balance),
        }
        if c_id in contractor_rows:
            contractor_rows[c_id].append(row)

    # ─── ۵. ساخت فایل اکسل - یک شیت به ازای هر پیمانکار ───────────────────
    wb = Workbook()
    _register_named_styles(wb)
    wb.remove(wb.active)  # حذف شیت پیش‌فرض خالی

    for c_id in sorted(contractor_rows.keys()):
        contractor = contractors.get(c_id)
        if not contractor:
            continue

        rows = contractor_rows[c_id]
        sheet_name = contractor.get_full_name()[:31]  # حداکثر ۳۱ کاراکتر
        ws = wb.create_sheet(title=sheet_name)

        _build_contractor_sheet(ws, contractor, rows)
        
        if not is_superuser:
            ws.protection.sheet = True
            ws.protection.password = "jahanpars2026"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_global_material_balance_rows_data() -> list[dict]:
    """
    دریافت اطلاعات موازنه متریال کل کارگاه به صورت تجمیعی.
    """
    from .models import WarehouseTransaction, TechnicalOfficeApproval, MaterialItem, Contractor
    from django.db.models import Sum
    from decimal import Decimal, ROUND_HALF_UP

    # ─── ۱. بارگذاری و جمع‌بندی داده‌های تراکنش خروجی انبار کل ─────────────────
    issue_qs = WarehouseTransaction.objects.filter(transaction_type='OUT')
    issued_aggs = issue_qs.values('contractor_id', 'material_id', 'contract_number', 'contract_subject').annotate(total_qty=Sum('quantity'))
    issued_map: dict[tuple, Decimal] = {}
    for agg in issued_aggs:
        k = (agg['contractor_id'], agg['material_id'], agg['contract_number'] or '', agg['contract_subject'] or '')
        issued_map[k] = agg['total_qty'] or Decimal('0')

    approval_qs = TechnicalOfficeApproval.objects.all()
    appr_aggs = approval_qs.values('contractor_id', 'material_id', 'contract_number', 'contract_subject').annotate(total_appr=Sum('approved_quantity'))
    approved_map: dict[tuple, Decimal] = {}
    for agg in appr_aggs:
        k = (agg['contractor_id'], agg['material_id'], agg['contract_number'] or '', agg['contract_subject'] or '')
        approved_map[k] = agg['total_appr'] or Decimal('0')

    # ─── ۳. جمع‌آوری تمام متریال‌های استفاده شده ─────────────────────────────
    all_keys = set(issued_map.keys()) | set(approved_map.keys())

    if not all_keys:
        return []

    contractor_ids = {k[0] for k in all_keys if k[0] is not None}
    material_ids   = {k[1] for k in all_keys}
    
    contractors = {c.id: c for c in Contractor.objects.filter(id__in=contractor_ids)}
    materials = {m.id: m for m in MaterialItem.objects.select_related('work_category').filter(id__in=material_ids)}

    # ─── ۴. آماده‌سازی ردیف‌های گزارش ─────────────────────────────
    rows = []
    
    # مرتب‌سازی بر اساس نام پیمانکار، سپس نام کالا
    def get_sort_key(k):
        c_id, m_id, c_num, c_subj = k
        contractor = contractors.get(c_id)
        mat = materials.get(m_id)
        c_name = contractor.get_full_name() if contractor else ""
        m_name = mat.name if mat else ""
        return (c_name, m_name, c_num, c_subj)
        
    sorted_keys = sorted(all_keys, key=get_sort_key)
    
    for c_id, m_id, contract_num, contract_subj in sorted_keys:
        material = materials.get(m_id)
        contractor = contractors.get(c_id)
        if not material or not contractor:
            continue

        total_issued  = issued_map.get((c_id, m_id, contract_num, contract_subj), Decimal('0'))
        
        # بررسی اینکه آیا اصلا تاییدیه‌ای برای این قرارداد و متریال وجود دارد یا خیر
        if (c_id, m_id, contract_num, contract_subj) not in approved_map:
            approved_work = Decimal('0')
            allowed_waste = Decimal('0')
            balance = "در دست بررسی برای تایید دفتر فنی"
        else:
            approved_work = approved_map[(c_id, m_id, contract_num, contract_subj)]
            waste_pct     = material.waste_percentage / Decimal('100')
            allowed_waste = (approved_work * waste_pct).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            balance       = (total_issued - (approved_work + allowed_waste)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        row = {
            'contractor_name': contractor.get_full_name(),
            'contract_number': contract_num or "—",
            'contract_subject': contract_subj or "—",
            'work_category':  material.work_category.name if material.work_category else "—",
            'material_name':  material.name,
            'size':           material.size or "—",
            'mat_type':       material.material_type or "—",
            'thickness':      material.thickness or "—",
            'unit':           material.get_unit_display(),
            'total_issued':   float(total_issued),
            'approved_work':  float(approved_work),
            'waste_pct':      float(material.waste_percentage),
            'allowed_waste':  float(allowed_waste),
            'balance':        float(balance) if isinstance(balance, Decimal) else balance,
            'balance_label':  _balance_label(balance),
        }
        rows.append(row)

    return rows


def generate_global_material_balance_excel(is_superuser: bool = False) -> bytes:
    """
    تولید گزارش موازنه متریال کل کارگاه به صورت تجمیعی.
    """
    rows = get_global_material_balance_rows_data()
    
    if not rows:
        # فایل خالی
        wb = Workbook()
        ws = wb.active
        ws.title = "بدون داده"
        ws.sheet_view.rightToLeft = True
        ws.merge_cells('A1:E1')
        c = ws['A1']
        c.value = "هیچ داده‌ای برای گزارش‌گیری وجود ندارد."
        c.font = Font(name='Calibri', size=12, bold=True, color="9C0006")
        c.alignment = _rtl_alignment(horizontal='center')
        if not is_superuser:
            ws.protection.sheet = True
            ws.protection.password = "jahanpars2026"
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ─── ۵. ساخت فایل اکسل ───────────────────
    wb = Workbook()
    _register_named_styles(wb)
    wb.remove(wb.active)  # حذف شیت پیش‌فرض خالی
    
    ws = wb.create_sheet(title="موازنه کل کارگاه")
    _build_global_sheet(ws, rows)
    
    if not is_superuser:
        ws.protection.sheet = True
        ws.protection.password = "jahanpars2026"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# تابع کمکی: دانلود مستقیم از Django View
# ─────────────────────────────────────────────────────────────────────────────
def get_balance_excel_response(contractor_id=None, material_id=None, is_superuser=False):
    """
    یک HttpResponse آماده برای دانلود فایل اکسل از Django View برمی‌گرداند.

    مثال استفاده در views.py:
        from .services import get_balance_excel_response

        def download_balance(request):
            contractor_id = request.GET.get('contractor')
            return get_balance_excel_response(contractor_id=contractor_id, is_superuser=request.user.is_superuser)
    """
    from django.http import HttpResponse

    file_bytes = generate_material_balance_excel(
        contractor_id=contractor_id,
        material_id=material_id,
        is_superuser=is_superuser,
    )
    filename = f"material_balance_{jdatetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        content=file_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def get_global_balance_excel_response(is_superuser=False):
    """
    دانلود گزارش کلی موازنه.
    """
    from django.http import HttpResponse

    file_bytes = generate_global_material_balance_excel(is_superuser=is_superuser)
    filename = f"global_material_balance_{jdatetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        content=file_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# تابع خروجی JSON برای API داشبورد
# ─────────────────────────────────────────────────────────────────────────────
def get_contractors_balance_summary() -> list[dict]:
    """
    محاسبه موازنه کلی تمامی پیمانکاران برای نمایش در داشبورد یا API.
    برگرداندن لیستی از دیکشنری‌ها.
    """
    from .models import WarehouseTransaction, TechnicalOfficeApproval, MaterialItem, Contractor
    from django.db.models import Sum

    # تراکنش‌های خروجی
    issue_qs = WarehouseTransaction.objects.filter(transaction_type='OUT')
    issued_aggs = issue_qs.values('contractor_id', 'material_id', 'contract_number', 'contract_subject').annotate(total_qty=Sum('quantity'))
    issued_map: dict[tuple, Decimal] = {}
    for agg in issued_aggs:
        k = (agg['contractor_id'], agg['material_id'], agg['contract_number'] or '', agg['contract_subject'] or '')
        issued_map[k] = agg['total_qty'] or Decimal('0')

    # تاییدیه‌ها
    approval_qs = TechnicalOfficeApproval.objects.all()
    appr_aggs = approval_qs.values('contractor_id', 'material_id', 'contract_number', 'contract_subject').annotate(total_appr=Sum('approved_quantity'))
    approved_map: dict[tuple, Decimal] = {}
    for agg in appr_aggs:
        k = (agg['contractor_id'], agg['material_id'], agg['contract_number'] or '', agg['contract_subject'] or '')
        approved_map[k] = agg['total_appr'] or Decimal('0')

    # کلیدها و مدل‌ها
    all_keys = set(issued_map.keys()) | set(approved_map.keys())
    contractor_ids = {k[0] for k in all_keys if k[0] is not None}
    material_ids   = {k[1] for k in all_keys}
    
    contractors = {c.id: c for c in Contractor.objects.filter(id__in=contractor_ids)}
    materials = {m.id: m for m in MaterialItem.objects.filter(id__in=material_ids)}

    # محاسبه موازنه برای هر ترکیب (پیمانکار، متریال) و جمع زدن بر اساس پیمانکار به تفکیک واحد
    contractor_balances = {c_id: {} for c_id in contractor_ids}
    contractor_under_review = {c_id: 0 for c_id in contractor_ids}

    for c_id, m_id, c_num, c_subj in all_keys:
        material = materials.get(m_id)
        contractor = contractors.get(c_id)
        if not material or not contractor:
            continue

        total_issued  = issued_map.get((c_id, m_id, c_num, c_subj), Decimal('0'))
        
        if (c_id, m_id, c_num, c_subj) not in approved_map:
            # اگر تاییدیه‌ای وجود ندارد، در دست بررسی است و در مجموع بالانس تاثیر نمی‌دهد
            contractor_under_review[c_id] += 1
        else:
            approved_work = approved_map[(c_id, m_id, c_num, c_subj)]
            waste_pct     = material.waste_percentage / Decimal('100')
            allowed_waste = (approved_work * waste_pct).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            balance       = (total_issued - (approved_work + allowed_waste)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            
            unit = material.unit
            if unit not in contractor_balances[c_id]:
                contractor_balances[c_id][unit] = Decimal('0')
            contractor_balances[c_id][unit] += balance

    # ساخت خروجی نهایی به شکل لیست
    summary = []
    for c_id, unit_bals in contractor_balances.items():
        contractor = contractors.get(c_id)
        balances_json = {unit: float(val) for unit, val in unit_bals.items()}
        summary.append({
            'contractor_id': c_id,
            'contractor_name': contractor.get_full_name() if contractor else "ناشناس",
            'balances': balances_json,
            'under_review_count': contractor_under_review[c_id],
        })

    # مرتب‌سازی بر اساس نام پیمانکار
    summary.sort(key=lambda x: x['contractor_name'])
    return summary


# ─────────────────────────────────────────────────────────────────────────────
# خروجی اکسل انباردار
# ─────────────────────────────────────────────────────────────────────────────
def generate_warehouse_inventory_excel(is_superuser: bool = False) -> bytes:
    # تولید فایل اکسل از لیست تمامی تراکنش‌های انبار.
    from .models import WarehouseTransaction
    qs = WarehouseTransaction.objects.select_related('contractor', 'material', 'material__work_category').order_by('-date', '-created_at')

    wb = Workbook()
    _register_named_styles(wb)
    ws = wb.active
    ws.title = "تراکنش‌های انبار"
    ws.sheet_view.rightToLeft = True

    TOTAL_COLS = 14
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    _apply_header_style(ws.cell(1, 1), "لیست تراکنش‌های انبار - شرکت جهانپارس", font_size=16)
    ws.row_dimensions[1].height = 38

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    sub = ws.cell(2, 1)
    sub.value = f"تاریخ گزارش: {jdatetime.date.today().isoformat()}"
    sub.font = Font(name='Calibri', size=10)
    sub.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_BG)
    sub.alignment = _rtl_alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=TOTAL_COLS)
    ws.cell(3, 1).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
    ws.row_dimensions[3].height = 6

    HEADERS = ["ردیف", "نوع تراکنش", "تاریخ", "پیمانکار", "شماره قرارداد", "موضوع قرارداد", "رسته کاری", "کالا", "سایز", "جنس", "ضخامت", "تعداد/مقدار", "واحد", "توضیحات"]
    COL_WIDTHS = [6, 12, 14, 20, 16, 25, 16, 24, 12, 14, 12, 14, 10, 30]

    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(4, i)
        _apply_header_style(c, h, font_size=10, bg_color=COLOR_SUB_HEADER_BG)
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[4].height = 28

    last_row = 4
    for row_idx, txn in enumerate(qs.iterator(chunk_size=2000)):
        r = row_idx + 5
        last_row = r
        _apply_data_style(ws.cell(r, 1), row_idx + 1, row_idx, horizontal='center')
        
        type_cell = ws.cell(r, 2)
        type_cell.value = "ورود" if txn.transaction_type == 'IN' else "خروج"
        type_cell.style = "data_in" if txn.transaction_type == 'IN' else "data_out"

        _apply_data_style(ws.cell(r, 3), str(txn.date), row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 4), txn.contractor.get_full_name() if txn.contractor else "—", row_idx)
        _apply_data_style(ws.cell(r, 5), txn.contract_number or "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 6), txn.contract_subject or "—", row_idx)
        _apply_data_style(ws.cell(r, 7), txn.material.work_category.name if txn.material and txn.material.work_category else "—", row_idx)
        _apply_data_style(ws.cell(r, 8), txn.material.name if txn.material else "—", row_idx)
        _apply_data_style(ws.cell(r, 9), txn.material.size if txn.material and txn.material.size else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 10), txn.material.material_type if txn.material and txn.material.material_type else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 11), txn.material.thickness if txn.material and txn.material.thickness else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 12), float(txn.quantity), row_idx, number_format='#,##0.00')
        _apply_data_style(ws.cell(r, 13), txn.material.get_unit_display() if txn.material else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 14), txn.bill_of_lading or "", row_idx)

    ws.auto_filter.ref = f"A4:N{last_row}"
    ws.freeze_panes = "A5"

    if not is_superuser:
        ws.protection.sheet = True
        ws.protection.password = "jahanpars2026"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_warehouse_inventory_excel_response(is_superuser: bool = False):
    from django.http import HttpResponse
    file_bytes = generate_warehouse_inventory_excel(is_superuser=is_superuser)
    filename = f"warehouse_inventory_{jdatetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        content=file_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# خروجی اکسل لیست پیمانکاران
# ─────────────────────────────────────────────────────────────────────────────
def generate_contractors_list_excel(is_superuser: bool = False) -> bytes:
    from .models import Contractor
    qs = Contractor.objects.all().order_by('first_name', 'last_name')

    wb = Workbook()
    ws = wb.active
    ws.title = "لیست پیمانکاران"
    ws.sheet_view.rightToLeft = True

    TOTAL_COLS = 3
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    _apply_header_style(ws.cell(1, 1), "لیست پیمانکاران - شرکت جهانپارس", font_size=16)
    ws.row_dimensions[1].height = 38

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    sub = ws.cell(2, 1)
    sub.value = f"تاریخ گزارش: {jdatetime.date.today().isoformat()}"
    sub.font = Font(name='Calibri', size=10)
    sub.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_BG)
    sub.alignment = _rtl_alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=TOTAL_COLS)
    ws.cell(3, 1).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
    ws.row_dimensions[3].height = 6

    HEADERS = ["ردیف", "نام", "نام خانوادگی"]
    COL_WIDTHS = [10, 30, 30]

    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(4, i)
        _apply_header_style(c, h, font_size=10, bg_color=COLOR_SUB_HEADER_BG)
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[4].height = 28

    last_row = 4
    for row_idx, contractor in enumerate(qs.iterator(chunk_size=2000)):
        r = row_idx + 5
        last_row = r
        _apply_data_style(ws.cell(r, 1), row_idx + 1, row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 2), contractor.first_name, row_idx)
        _apply_data_style(ws.cell(r, 3), contractor.last_name, row_idx)

    ws.auto_filter.ref = f"A4:C{last_row}"
    ws.freeze_panes = "A5"

    if not is_superuser:
        ws.protection.sheet = True
        ws.protection.password = "jahanpars2026"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_contractors_excel_response(is_superuser: bool = False):
    from django.http import HttpResponse
    file_bytes = generate_contractors_list_excel(is_superuser=is_superuser)
    filename = f"contractors_list_{jdatetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        content=file_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# خروجی اکسل لیست تاییدیه‌ها
# ─────────────────────────────────────────────────────────────────────────────
def generate_approvals_list_excel(is_superuser: bool = False) -> bytes:
    from .models import TechnicalOfficeApproval
    qs = TechnicalOfficeApproval.objects.select_related('contractor', 'material').order_by('-approval_date', '-created_at')

    wb = Workbook()
    _register_named_styles(wb)
    ws = wb.active
    ws.title = "لیست تاییدیه‌های دفتر فنی"
    ws.sheet_view.rightToLeft = True

    TOTAL_COLS = 12
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    _apply_header_style(ws.cell(1, 1), "لیست تاییدیه‌های دفتر فنی - شرکت جهانپارس", font_size=16)
    ws.row_dimensions[1].height = 38

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLS)
    sub = ws.cell(2, 1)
    sub.value = f"تاریخ گزارش: {jdatetime.date.today().isoformat()}"
    sub.font = Font(name='Calibri', size=10)
    sub.fill = PatternFill(fill_type='solid', fgColor=COLOR_INFO_BG)
    sub.alignment = _rtl_alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=TOTAL_COLS)
    ws.cell(3, 1).fill = PatternFill(fill_type='solid', fgColor=COLOR_HEADER_BG)
    ws.row_dimensions[3].height = 6

    HEADERS = ["ردیف", "تاریخ تایید", "پیمانکار", "نام کالا", "سایز", "جنس", "ضخامت", "واحد", "موضوع / توضیحات", "مقدار تایید شده", "شماره قرارداد", "توضیحات تکمیلی"]
    COL_WIDTHS = [6, 14, 25, 24, 12, 14, 12, 10, 30, 15, 15, 30]

    for i, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(4, i)
        _apply_header_style(c, h, font_size=10, bg_color=COLOR_SUB_HEADER_BG)
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[4].height = 28

    last_row = 4
    for row_idx, approval in enumerate(qs.iterator(chunk_size=2000)):
        r = row_idx + 5
        last_row = r
        _apply_data_style(ws.cell(r, 1), row_idx + 1, row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 2), str(approval.approval_date), row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 3), approval.contractor.get_full_name() if approval.contractor else "—", row_idx)
        _apply_data_style(ws.cell(r, 4), approval.material.name if approval.material else "—", row_idx)
        _apply_data_style(ws.cell(r, 5), approval.material.size if approval.material and approval.material.size else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 6), approval.material.material_type if approval.material and approval.material.material_type else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 7), approval.material.thickness if approval.material and approval.material.thickness else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 8), approval.material.get_unit_display() if approval.material else "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 9), approval.contract_subject or "—", row_idx)
        _apply_data_style(ws.cell(r, 10), float(approval.approved_quantity), row_idx, number_format='#,##0.00')
        _apply_data_style(ws.cell(r, 11), approval.contract_number or "—", row_idx, horizontal='center')
        _apply_data_style(ws.cell(r, 12), "—", row_idx)

    ws.auto_filter.ref = f"A4:L{last_row}"
    ws.freeze_panes = "A5"

    if not is_superuser:
        ws.protection.sheet = True
        ws.protection.password = "jahanpars2026"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def get_approvals_excel_response(is_superuser: bool = False):
    from django.http import HttpResponse
    file_bytes = generate_approvals_list_excel(is_superuser=is_superuser)
    filename = f"approvals_list_{jdatetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        content=file_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
